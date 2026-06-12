"""One-off importer: loads 'Room bookings 2025-2026.xlsx' into the database.

Layout it expects (as found in the workbook):
  - one sheet per week; day blocks start where column A or T holds
    "Monday 13/10/25" etc.; the header row also holds the slot times
  - rows below a day header are rooms until a blank/another header
  - cell values are the course/activity; consecutive identical values
    merge into one booking; Excel comments go into notes, with the
    comment author recorded as booked_by

Usage:  python import_spreadsheet.py [--replace]
        --replace wipes existing bookings/rooms/series first (asks to confirm)
"""
import re
import sys
from datetime import date, datetime, time, timedelta

import openpyxl

from app import create_app
from app.models import AuditLog, Booking, BookingSeries, Room, audit, db

WORKBOOK = "Room bookings 2025-2026.xlsx"
DAY_RE = re.compile(
    r"(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\s+"
    r"(\d{1,2})[/.](\d{1,2})[/.](\d{2,4})", re.I)
ROOM_CAP_RE = re.compile(r"^(.*?)\s*\((\d+)\)\s*$")

# The workbook spells some rooms inconsistently — map (lowercased) variants
# to one canonical name.
ROOM_ALIASES = {
    "nuffield": "Nuffield",
    "nicksonnuffield": "Nuffield",  # typo on '29th Sept' Friday block
    "joint msc lecture room": "Joint Masters Lecture Room",
    "collabrative learning room(clr)": "Collaborative Learning Room (CLR)",
    "collobrative learning room(clr)": "Collaborative Learning Room (CLR)",
    "collaborative learning room": "Collaborative Learning Room (CLR)",
}
LAST_SLOT = time(17, 0)
LAST_SLOT_END = time(18, 0)


def parse_day_header(value):
    if not isinstance(value, str):
        return None
    m = DAY_RE.search(value)
    if not m:
        return None
    d, mth, y = int(m.group(2)), int(m.group(3)), int(m.group(4))
    if y < 100:
        y += 2000
    try:
        return date(y, mth, d)
    except ValueError:
        return None


def parse_slot(value):
    """Header cell -> slot start time (slots are 30 min; '17:00+' is last)."""
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        s = value.strip().rstrip("+")
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                pass
    return None


def get_room(name, order_hint):
    name = " ".join(str(name).split())
    capacity = None
    m = ROOM_CAP_RE.match(name)
    if m:
        name, capacity = m.group(1).strip(), int(m.group(2))
    name = ROOM_ALIASES.get(name.lower(), name)
    room = Room.query.filter_by(name=name).first()
    if not room:
        room = Room(name=name, capacity=capacity, sort_order=order_hint)
        db.session.add(room)
        db.session.flush()
    elif capacity and not room.capacity:
        room.capacity = capacity
    return room


def slot_end(slot):
    if slot == LAST_SLOT:
        return LAST_SLOT_END
    return (datetime.combine(date.min, slot) + timedelta(minutes=30)).time()


def import_sheet(ws, stats):
    # Find day-header cells in columns A (1) and T (20).
    headers = []
    for row in ws.iter_rows():
        for cell in (row[0], row[19] if len(row) > 19 else None):
            if cell is not None and (d := parse_day_header(cell.value)):
                headers.append((cell.row, cell.column, d))
    for header_row, header_col, day in headers:
        # Slot times sit to the right of the day header on the same row.
        slot_cols = {}
        for c in range(header_col + 1, header_col + 18):
            if (s := parse_slot(ws.cell(header_row, c).value)) is not None:
                slot_cols[c] = s
        if not slot_cols:
            continue
        # Room rows follow until a blank cell or the next day header.
        r = header_row + 1
        while r <= ws.max_row:
            name_cell = ws.cell(r, header_col).value
            if name_cell is None or str(name_cell).strip() == "" \
                    or parse_day_header(name_cell):
                break
            room = get_room(name_cell, order_hint=r - header_row)
            run = None  # (activity, start, end, notes, booked_by)
            for c in sorted(slot_cols):
                cell = ws.cell(r, c)
                val = str(cell.value).strip() if cell.value is not None else None
                note, author = None, None
                if cell.comment and cell.comment.text:
                    text = cell.comment.text.strip()
                    if ":" in text.split("\n")[0]:
                        author, note = text.split(":", 1)
                        author, note = author.strip(), note.strip()
                    else:
                        note = text
                if run and val == run["activity"]:
                    run["end"] = slot_end(slot_cols[c])
                    if note and note not in (run["notes"] or ""):
                        run["notes"] = f"{run['notes']}; {note}" if run["notes"] else note
                    run["booked_by"] = run["booked_by"] or author
                else:
                    if run:
                        save_run(room, day, run, stats)
                    run = None
                    if val:
                        run = {"activity": val, "start": slot_cols[c],
                               "end": slot_end(slot_cols[c]),
                               "notes": note, "booked_by": author}
            if run:
                save_run(room, day, run, stats)
            r += 1


def save_run(room, day, run, stats):
    existing = Booking.query.filter_by(
        room_id=room.id, date=day,
        start_time=run["start"], end_time=run["end"],
        activity=run["activity"]).first()
    if existing:
        stats["duplicates"] += 1
        return
    db.session.add(Booking(
        room_id=room.id, date=day,
        start_time=run["start"], end_time=run["end"],
        activity=run["activity"], notes=run["notes"],
        booked_by=run["booked_by"] or "Imported from spreadsheet",
    ))
    stats["bookings"] += 1


def main():
    app = create_app()
    with app.app_context():
        if "--replace" in sys.argv:
            counts = (Booking.query.count(), Room.query.count())
            answer = input(f"Delete {counts[0]} bookings and {counts[1]} rooms "
                           "before importing? [y/N] ")
            if answer.lower() != "y":
                print("Aborted.")
                return
            Booking.query.delete()
            BookingSeries.query.delete()
            Room.query.delete()
            db.session.commit()

        wb = openpyxl.load_workbook(WORKBOOK)
        stats = {"bookings": 0, "duplicates": 0}
        for name in wb.sheetnames:
            before = stats["bookings"]
            import_sheet(wb[name], stats)
            print(f"  {name}: {stats['bookings'] - before} bookings")
        audit("importer", "import", None,
              {"workbook": WORKBOOK, **stats, "when": datetime.utcnow()})
        db.session.commit()
        print(f"\nDone: {stats['bookings']} bookings imported, "
              f"{stats['duplicates']} duplicates skipped.")
        print(f"Rooms: {[r.name for r in Room.query.order_by(Room.sort_order)]}")


if __name__ == "__main__":
    main()
